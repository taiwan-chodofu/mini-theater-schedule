"""
都内近郊ミニシアター上映スケジュール取得スクリプト

対象映画館:
  イメージフォーラム、下高井戸シネマ、目黒シネマ、早稲田松竹、
  キネカ大森、ストレンジャー、ユーロスペース、K's cinema、
  武蔵野館、ジャック＆ベティ

出力: Excelファイル（日付付き）
実行: py theater_schedule.py
"""

import requests
from bs4 import BeautifulSoup
from datetime import datetime, date, timezone, timedelta
import re
import json
import os
import html
try:
    from playwright.sync_api import sync_playwright
    HAS_PLAYWRIGHT = True
except ImportError:
    HAS_PLAYWRIGHT = False

# --- 映画館マスタデータ ---
THEATERS = {
    "シアター・イメージフォーラム": {
        "eigacom_url": "https://eiga.com/theater/13/130301/3028/",
        "url_top": "https://www.imageforum.co.jp/theatre/",
        "station": "渋谷駅（東口 徒歩8分）",
        "address": "東京都渋谷区渋谷2-10-2",
        "lat": 35.66037,
        "lng": 139.707047,
    },
    "下高井戸シネマ": {
        "eigacom_url": "https://eiga.com/theater/13/130606/3064/",
        "url_top": "https://www.shimotakaidocinema.com/",
        "station": "下高井戸駅（京王線・東急世田谷線 徒歩2分）",
        "address": "東京都世田谷区松原3-27-26",
        "lat": 35.665997,
        "lng": 139.642639,
    },
    "目黒シネマ": {
        "eigacom_url": "https://eiga.com/theater/13/130609/3069/",
        "url_top": "http://www.okura-movie.co.jp/meguro_cinema/",
        "station": "目黒駅（JR・東急目黒線 徒歩3分）",
        "address": "東京都品川区上大崎2-24-15 目黒西口ビルB1",
        "lat": 35.634987,
        "lng": 139.714874,
    },
    "早稲田松竹": {
        "eigacom_url": "https://eiga.com/theater/13/130611/3071/",
        "url_top": "https://wasedashochiku.co.jp/",
        "station": "高田馬場駅（JR・西武新宿線・東西線 徒歩5分）",
        "address": "東京都新宿区高田馬場1-5-16",
        "lat": 35.711731,
        "lng": 139.708176,
    },
    "キネカ大森": {
        "eigacom_url": "https://eiga.com/theater/13/130713/3095/",
        "url_top": "https://www.ttcg.jp/cineka_omori/",
        "station": "大森駅（JR京浜東北線 徒歩1分）",
        "address": "東京都品川区南大井6-27-25 西友大森店5階",
        "lat": 35.588146,
        "lng": 139.729919,
    },
    "ストレンジャー": {
        "eigacom_url": "https://eiga.com/theater/13/130718/3319/",
        "url_top": "https://stranger.jp/",
        "station": "菊川駅（都営新宿線 徒歩1分）",
        "address": "東京都墨田区菊川3-7-1 菊川会館ビル1F",
        "lat": 35.688,
        "lng": 139.806183,
    },
    "ユーロスペース": {
        "eigacom_url": "https://eiga.com/theater/13/130301/3044/",
        "url_top": "https://www.eurospace.co.jp/",
        "station": "渋谷駅（道玄坂方面 徒歩10分）",
        "address": "東京都渋谷区円山町1-5 KINOHAUS 2F",
        "lat": 35.659393,
        "lng": 139.695404,
    },
    "K's cinema": {
        "eigacom_url": "https://eiga.com/theater/13/130201/3018/",
        "url_top": "https://www.ks-cinema.com/",
        "station": "新宿駅（東口 徒歩3分）",
        "address": "東京都新宿区新宿3-35-13 3F",
        "lat": 35.69022,
        "lng": 139.702652,
    },
    "新宿武蔵野館": {
        "eigacom_url": "https://eiga.com/theater/13/130201/3026/",
        "url_top": "https://shinjuku.musashino-k.jp/",
        "station": "新宿駅（東口 徒歩2分）",
        "address": "東京都新宿区新宿3-27-10 武蔵野ビル3F",
        "lat": 35.691216,
        "lng": 139.702286,
    },
    "シネマ・ジャック＆ベティ": {
        "eigacom_url": "https://eiga.com/theater/14/140108/3247/",
        "url_top": "https://www.jackandbetty.net/",
        "station": "黄金町駅（京急線 徒歩5分）",
        "address": "神奈川県横浜市中区若葉町3-51",
        "lat": 35.44046,
        "lng": 139.626511,
    },
    # --- 千代田区 ---
    "神保町シアター": {
        "eigacom_url": "https://eiga.com/theater/13/130710/3302/",
        "url_top": "https://www.shogakukan.co.jp/jinbocho-theater/",
        "station": "神保町駅（都営三田線・新宿線・半蔵門線 徒歩3分）",
        "address": "東京都千代田区神田神保町1-23",
        "lat": 35.695042,
        "lng": 139.760986,
    },
    "ヒューマントラストシネマ有楽町": {
        "eigacom_url": "https://eiga.com/theater/13/130102/3004/",
        "url_top": "https://ttcg.jp/human_yurakucho/",
        "station": "有楽町駅（JR 徒歩1分）",
        "address": "東京都千代田区有楽町2-7-1 イトシアプラザ4F",
        "lat": 35.674213,
        "lng": 139.763657,
    },
    "角川シネマ有楽町": {
        "eigacom_url": "https://eiga.com/theater/13/130102/3248/",
        "url_top": "https://www.kadokawa-cinema.jp/yurakucho/",
        "station": "有楽町駅（JR 徒歩1分）",
        "address": "東京都千代田区有楽町1-11-1 読売会館8F",
        "lat": 35.675388,
        "lng": 139.762878,
    },
    "TOHOシネマズ シャンテ": {
        "eigacom_url": "https://eiga.com/theater/13/130102/3006/",
        "url_top": "https://hlo.tohotheater.jp/net/schedule/009/TNPI2000J01.do",
        "station": "日比谷駅（東京メトロ 徒歩2分）",
        "address": "東京都千代田区有楽町1-2-2",
        "lat": 35.672913,
        "lng": 139.759964,
    },
    "TOHOシネマズ 日比谷": {
        "eigacom_url": "https://eiga.com/theater/13/130102/3281/",
        "url_top": "https://hlo.tohotheater.jp/net/schedule/076/TNPI2000J01.do",
        "station": "日比谷駅（東京メトロ 直結）",
        "address": "東京都千代田区有楽町1-1-2 東京ミッドタウン日比谷4F",
        "lat": 35.674088,
        "lng": 139.759552,
    },
    "アテネ・フランセ文化センター": {
        "eigacom_url": "https://eiga.com/theater/13/130614/3299/",
        "url_top": "http://www.athenee.net/culturalcenter/",
        "station": "御茶ノ水駅（JR 徒歩5分）",
        "address": "東京都千代田区神田駿河台2-11 アテネ・フランセ4F",
        "lat": 35.700912,
        "lng": 139.758102,
    },
    "シネマリス": {
        "eigacom_url": "https://eiga.com/theater/13/130710/3331/",
        "url_top": "https://www.cinemairis.com/",
        "station": "神保町駅（都営三田線・新宿線・半蔵門線 徒歩5分）",
        "address": "東京都千代田区神田小川町3-14-3 ilusa B1F",
        "lat": 35.696548,
        "lng": 139.760818,
    },
    # --- 中央区 ---
    "シネスイッチ銀座": {
        "eigacom_url": "https://eiga.com/theater/13/130101/3005/",
        "url_top": "https://www.cineswitch.com/",
        "station": "銀座駅（東京メトロ 徒歩1分）",
        "address": "東京都中央区銀座4-4-5",
        "lat": 35.672066,
        "lng": 139.765121,
    },
    "国立映画アーカイブ": {
        "eigacom_url": "https://eiga.com/theater/13/130101/3300/",
        "url_top": "https://www.nfaj.go.jp/",
        "station": "京橋駅（東京メトロ 徒歩1分）",
        "address": "東京都中央区京橋3-7-6",
        "lat": 35.675701,
        "lng": 139.770599,
    },
    # --- 新宿区（追加分） ---
    "シネマート新宿": {
        "eigacom_url": "https://eiga.com/theater/13/130201/3020/",
        "url_top": "https://www.cinemart.co.jp/theater/shinjuku/",
        "station": "新宿三丁目駅（東京メトロ 徒歩3分）",
        "address": "東京都新宿区新宿3-13-3 新宿文化ビル6F・7F",
        "lat": 35.691875,
        "lng": 139.705612,
    },
    "テアトル新宿": {
        "eigacom_url": "https://eiga.com/theater/13/130201/3022/",
        "url_top": "https://ttcg.jp/theatre_shinjuku/",
        "station": "新宿三丁目駅（東京メトロ 徒歩3分）",
        "address": "東京都新宿区新宿3-14-20 新宿テアトルビルB1F",
        "lat": 35.692406,
        "lng": 139.705215,
    },
    "アンスティチュ・フランセ東京": {
        "eigacom_url": "https://eiga.com/theater/13/130604/3310/",
        "url_top": "https://www.institutfrancais.jp/tokyo/",
        "station": "飯田橋駅（JR・東京メトロ 徒歩7分）",
        "address": "東京都新宿区市谷船河原町15",
        "lat": 35.698498,
        "lng": 139.739944,
    },
    # --- 目黒区 ---
    "東京都写真美術館ホール": {
        "eigacom_url": "https://eiga.com/theater/13/130608/3067/",
        "url_top": "https://topmuseum.jp/",
        "station": "恵比寿駅（JR 徒歩7分）",
        "address": "東京都目黒区三田1-13-3 恵比寿ガーデンプレイス内",
        "lat": 35.642181,
        "lng": 139.715027,
    },
    # --- 世田谷区 ---
    "下北沢トリウッド": {
        "eigacom_url": "https://eiga.com/theater/13/130613/3277/",
        "url_top": "https://www.tollywood.jp/",
        "station": "下北沢駅（小田急線・京王井の頭線 徒歩5分）",
        "address": "東京都世田谷区代沢5-32-5 シェルボ下北沢2F",
        "lat": 35.658775,
        "lng": 139.667587,
    },
    "シモキタ-エキマエ-シネマ K2": {
        "eigacom_url": "https://eiga.com/theater/13/130613/3317/",
        "url_top": "https://k2.shimokita-ekimae.com/",
        "station": "下北沢駅（小田急線・京王井の頭線 徒歩1分）",
        "address": "東京都世田谷区北沢2-21-22 (tefu) lounge 2F",
        "lat": 35.660793,
        "lng": 139.666656,
    },
    # --- 渋谷区（追加分） ---
    "Bunkamuraル・シネマ 渋谷宮下": {
        "eigacom_url": "https://eiga.com/theater/13/130301/3321/",
        "url_top": "https://www.bunkamura.co.jp/cinema/",
        "station": "渋谷駅（JR 徒歩5分）",
        "address": "東京都渋谷区渋谷1-24-12 渋谷東映プラザ7F・9F",
        "lat": 35.659901,
        "lng": 139.702164,
    },
    "YEBISU GARDEN CINEMA": {
        "eigacom_url": "https://eiga.com/theater/13/130608/3261/",
        "url_top": "https://www.unitedcinemas.jp/ygc/",
        "station": "恵比寿駅（JR 徒歩5分）",
        "address": "東京都渋谷区恵比寿4-20-2 恵比寿ガーデンプレイス内",
        "lat": 35.643616,
        "lng": 139.71254,
    },
    "ヒューマントラストシネマ渋谷": {
        "eigacom_url": "https://eiga.com/theater/13/130301/3042/",
        "url_top": "https://ttcg.jp/human_shibuya/",
        "station": "渋谷駅（JR 徒歩5分）",
        "address": "東京都渋谷区渋谷1-23-16 ココチビル7・8F",
        "lat": 35.662025,
        "lng": 139.702377,
    },
    "シネマヴェーラ渋谷": {
        "eigacom_url": "https://eiga.com/theater/13/130301/3298/",
        "url_top": "http://www.cinemavera.com/",
        "station": "渋谷駅（道玄坂方面 徒歩10分）",
        "address": "東京都渋谷区円山町1-5 KINOHAUS 4F",
        "lat": 35.659393,
        "lng": 139.695404,
    },
    "渋谷シネクイント": {
        "eigacom_url": "https://eiga.com/theater/13/130301/3283/",
        "url_top": "https://www.cinequinto.com/",
        "station": "渋谷駅（ハチ公口 徒歩5分）",
        "address": "東京都渋谷区宇田川町20-11 渋谷三葉ビル7F",
        "lat": 35.660923,
        "lng": 139.699768,
    },
    "ホワイト シネクイント": {
        "eigacom_url": "https://eiga.com/theater/13/130301/3295/",
        "url_top": "https://white-cine-quinto.com/",
        "station": "渋谷駅（ハチ公口 徒歩5分）",
        "address": "東京都渋谷区宇田川町15-1 渋谷パルコ8F",
        "lat": 35.662212,
        "lng": 139.699203,
    },
    # --- 中野区 ---
    "ポレポレ東中野": {
        "eigacom_url": "https://eiga.com/theater/13/130612/3292/",
        "url_top": "https://pole2.co.jp/",
        "station": "東中野駅（JR 徒歩1分）",
        "address": "東京都中野区東中野4-4-1 ポレポレ坐ビルB1F",
        "lat": 35.706448,
        "lng": 139.684586,
    },
    # --- 豊島区 ---
    "池袋シネマ・ロサ": {
        "eigacom_url": "https://eiga.com/theater/13/130501/3052/",
        "url_top": "https://www.cinemarosa.net/",
        "station": "池袋駅（西口 徒歩2分）",
        "address": "東京都豊島区西池袋1-37-12 ロサ会館内",
        "lat": 35.732548,
        "lng": 139.709366,
    },
    "新文芸坐": {
        "eigacom_url": "https://eiga.com/theater/13/130501/3055/",
        "url_top": "https://www.shin-bungeiza.com/",
        "station": "池袋駅（東口 徒歩3分）",
        "address": "東京都豊島区東池袋1-43-5 マルハン池袋ビル3F",
        "lat": 35.732601,
        "lng": 139.713577,
    },
    "シネマハウス大塚": {
        "eigacom_url": "https://eiga.com/theater/13/130616/3311/",
        "url_top": "https://cinemahouse-otsuka.com/",
        "station": "大塚駅（JR 徒歩10分）",
        "address": "東京都豊島区巣鴨4-7-4-101",
        "lat": 35.735901,
        "lng": 139.730148,
    },
    # --- 北区 ---
    "CINEMA Chupki TABATA": {
        "eigacom_url": "https://eiga.com/theater/13/130717/3272/",
        "url_top": "https://chupki.jpn.org/",
        "station": "田端駅（JR 徒歩5分）",
        "address": "東京都北区東田端2-8-4",
        "lat": 35.741261,
        "lng": 139.761688,
    },
    # --- 足立区 ---
    "シネマブルースタジオ": {
        "eigacom_url": "https://eiga.com/theater/13/130715/3301/",
        "url_top": "https://www.cinema-st.com/",
        "station": "北千住駅（JR・東京メトロ 徒歩10分）",
        "address": "東京都足立区千住1-4-1 東京芸術センター2F",
        "lat": 35.746788,
        "lng": 139.80011,
    },
    # --- 品川区（追加） ---
    "T・ジョイPRINCE品川": {
        "eigacom_url": "https://eiga.com/theater/13/130703/3080/",
        "url_top": "https://tjoy.jp/tjoy_prince_shinagawa",
        "station": "品川駅（高輪口 徒歩2分）",
        "address": "東京都港区高輪4-10-30 品川プリンスホテル アネックスタワー3F",
        "lat": 35.628574,
        "lng": 139.736984,
    },
    "TOHOシネマズ 大井町": {
        "eigacom_url": "https://eiga.com/theater/13/130703/3333/",
        "url_top": "https://hlo.tohotheater.jp/net/schedule/085/TNPI2000J01.do",
        "station": "大井町駅（JR・りんかい線 徒歩2分）",
        "address": "東京都品川区大井1-2-1 阪急大井町ガーデン4F",
        "lat": 35.604897,
        "lng": 139.734665,
    },
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}



# JST固定（GitHub Actions runnerはUTCで動くため、date.today()だと
# JST早朝の実行時にUTC前日扱いになりeiga.comの日付ヘッダーと噛み合わない）
JST = timezone(timedelta(hours=9))


def today_jst():
    """JST基準の今日の日付を返す。"""
    return datetime.now(JST).date()

TODAY = today_jst()
# テスト用: 金曜シミュレート（確認後に戻す）
# TODAY = date(2026, 4, 17)
IS_WEEKDAY = TODAY.weekday() < 5  # 月〜金 = True

# 日本の祝日（2026年）
HOLIDAYS_2026 = {
    date(2026, 1, 1), date(2026, 1, 12), date(2026, 2, 11),
    date(2026, 2, 23), date(2026, 3, 20), date(2026, 4, 29),
    date(2026, 5, 3), date(2026, 5, 4), date(2026, 5, 5),
    date(2026, 5, 6), date(2026, 7, 20), date(2026, 8, 11),
    date(2026, 9, 21), date(2026, 9, 22), date(2026, 9, 23),
    date(2026, 10, 12), date(2026, 11, 3), date(2026, 11, 23),
}


def is_holiday(d):
    """土日祝かどうか判定。"""
    return d.weekday() >= 5 or d in HOLIDAYS_2026


def get_target_dates():
    """取得対象の日付リストを返す。金曜/祝前日は翌休日分も含む。"""
    from datetime import timedelta
    dates = [TODAY]
    # 翌日以降の連続する休日を追加
    next_day = TODAY + timedelta(days=1)
    while is_holiday(next_day):
        dates.append(next_day)
        next_day += timedelta(days=1)
    return dates


TARGET_DATES = get_target_dates()


def fetch_html(url, encoding=None, verify_ssl=True):
    """HTMLを取得して返す。失敗時はNone。"""
    try:
        resp = requests.get(
            url, headers=HEADERS, timeout=15, verify=verify_ssl
        )
        if encoding:
            resp.encoding = encoding
        else:
            resp.encoding = resp.apparent_encoding
        resp.raise_for_status()
        return resp.text
    except requests.exceptions.SSLError:
        # SSL問題時はverify=Falseでリトライ
        try:
            resp = requests.get(
                url, headers=HEADERS, timeout=15, verify=False
            )
            resp.encoding = resp.apparent_encoding
            resp.raise_for_status()
            return resp.text
        except Exception as e2:
            print(f"  [WARN] SSL回避後も取得失敗: {url} -> {e2}")
            return None
    except Exception as e:
        print(f"  [WARN] 取得失敗: {url} -> {e}")
        return None


def fetch_html_js(url, wait_ms=5000):
    """Playwrightでヘッドレスブラウザ経由でHTMLを取得。未インストール時はrequestsにフォールバック。"""
    if not HAS_PLAYWRIGHT:
        return fetch_html(url)
    for attempt in range(2):
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                page = browser.new_page()
                page.goto(url, timeout=45000)
                page.wait_for_timeout(wait_ms)
                html = page.content()
                browser.close()
                return html
        except Exception as e:
            if attempt == 0:
                print(f"  [WARN] リトライ中... ({e.__class__.__name__})")
            else:
                print(f"  [WARN] Playwright取得失敗: {url} -> {e}")
    return None


def extract_time(text):
    """テキストから上映開始時刻(HH:MM)を抽出。"""
    patterns = [
        r'(\d{1,2})[：:](\d{2})\s*[～〜~]',
        r'(\d{1,2})[：:](\d{2})\s*(?:～|〜|~|開始|上映)',
        r'(\d{1,2})[：:](\d{2})',
    ]
    times = []
    for pat in patterns:
        for m in re.finditer(pat, text):
            h, mi = int(m.group(1)), int(m.group(2))
            if 0 <= h <= 23 and 0 <= mi <= 59:
                times.append(f"{h:02d}:{mi:02d}")
    return times


def filter_by_time(showings):
    """平日は18:00-22:00開始のみ、休日は全件返す。"""
    if not IS_WEEKDAY:
        return showings
    filtered = []
    for s in showings:
        evening_times = []
        for t in s.get("times", []):
            h = int(t.split(":")[0])
            if 18 <= h <= 22:
                evening_times.append(t)
        if evening_times:
            filtered.append({**s, "times": evening_times})
    return filtered


# --- 個別パーサー ---

def parse_shimotakaido(html):
    """下高井戸シネマのスケジュールをパース。"""
    soup = BeautifulSoup(html, "html.parser")
    showings = []
    # 各作品ブロックを探す
    for block in soup.find_all("td"):
        text = block.get_text(separator="\n", strip=True)
        if not text or len(text) < 10:
            continue
        # 作品名候補: 太字やリンクテキスト
        title_tag = block.find(["a", "b", "strong"])
        title = title_tag.get_text(strip=True) if title_tag else ""
        if not title:
            lines = [l.strip() for l in text.split("\n") if l.strip()]
            title = lines[0] if lines else ""
        # 時刻抽出
        times = extract_time(text)
        # 概要: 監督・出演・あらすじ等
        desc_lines = []
        for line in text.split("\n"):
            line = line.strip()
            if any(kw in line for kw in ["監督", "出演", "年/", "年／"]):
                desc_lines.append(line)
            elif len(line) > 20 and "円" not in line and "：" not in line[:3]:
                desc_lines.append(line)
        desc = " ".join(desc_lines[:3])
        if title and times:
            showings.append({
                "title": title,
                "times": times,
                "description": desc[:200],
            })
    return showings


def parse_waseda(html):
    """早稲田松竹のスケジュールをパース。"""
    soup = BeautifulSoup(html, "html.parser")
    showings = []
    text = soup.get_text(separator="\n")
    # 作品名パターン: 日本語タイトル + 時刻
    blocks = re.split(r'\n(?=\d{1,2}/\d{1,2})', text)
    current_date_range = ""
    for block in blocks:
        lines = [l.strip() for l in block.split("\n") if l.strip()]
        if not lines:
            continue
        # 日付範囲を検出
        date_match = re.search(r'(\d{1,2}/\d{1,2})\s*[（(]\s*[日月火水木金土]', block)
        if date_match:
            current_date_range = lines[0] if lines else ""
        # 作品名と時刻のペアを探す
        for i, line in enumerate(lines):
            times = extract_time(line)
            if times:
                # 直前の行が作品名の可能性
                title_candidates = []
                for j in range(max(0, i - 3), i):
                    candidate = lines[j].strip()
                    if (candidate and not re.match(r'^\d', candidate)
                            and "円" not in candidate
                            and "チケット" not in candidate
                            and len(candidate) > 1):
                        title_candidates.append(candidate)
                if title_candidates:
                    title = title_candidates[-1]
                    # 監督情報を探す
                    desc = ""
                    for k in range(i + 1, min(len(lines), i + 5)):
                        if any(kw in lines[k] for kw in ["監督", "出演", "年/"]):
                            desc = lines[k]
                            break
                    showings.append({
                        "title": title,
                        "times": times,
                        "description": desc[:200],
                    })
    # 重複除去
    seen = set()
    unique = []
    for s in showings:
        key = s["title"]
        if key not in seen:
            seen.add(key)
            unique.append(s)
    return unique


def parse_jackandbetty(html):
    """ジャック＆ベティのスケジュールをパース。"""
    soup = BeautifulSoup(html, "html.parser")
    showings = []
    # 各作品ブロック
    for item in soup.select("div, td, li"):
        text = item.get_text(separator="\n", strip=True)
        if not text or len(text) < 20:
            continue
        # 時刻を含むブロックのみ
        times = extract_time(text)
        if not times:
            continue
        # 作品名: 最初の意味のある行
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        title = ""
        desc = ""
        for line in lines:
            if (not title and len(line) > 1
                    and not re.match(r'^[\d¥￥]', line)
                    and "終了日" not in line
                    and "〜" not in line[:5]):
                title = line
            if any(kw in line for kw in ["監督", "キャスト", "出演"]):
                desc += line + " "
        if title:
            showings.append({
                "title": title[:80],
                "times": times,
                "description": desc[:200].strip(),
            })
    # 重複除去
    seen = set()
    unique = []
    for s in showings:
        if s["title"] not in seen:
            seen.add(s["title"])
            unique.append(s)
    return unique


def parse_generic(html):
    """汎用パーサー: テキストから作品名と時刻を抽出。"""
    soup = BeautifulSoup(html, "html.parser")
    showings = []
    text = soup.get_text(separator="\n")
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    i = 0
    while i < len(lines):
        times = extract_time(lines[i])
        if times:
            # 直前の行を作品名として取得
            title = ""
            for j in range(max(0, i - 3), i):
                candidate = lines[j]
                if (len(candidate) > 1
                        and not re.match(r'^[\d¥￥]', candidate)
                        and "円" not in candidate):
                    title = candidate
            if not title and i > 0:
                title = lines[i - 1]
            # 直後の行から概要を取得
            desc = ""
            for k in range(i + 1, min(len(lines), i + 5)):
                if any(kw in lines[k] for kw in ["監督", "出演", "年/", "年／"]):
                    desc = lines[k]
                    break
                elif len(lines[k]) > 30:
                    desc = lines[k]
                    break
            if title:
                showings.append({
                    "title": title[:80],
                    "times": times,
                    "description": desc[:200],
                })
        i += 1
    # 重複除去
    seen = set()
    unique = []
    for s in showings:
        if s["title"] not in seen:
            seen.add(s["title"])
            unique.append(s)
    return unique


def parse_stranger(html):
    """ストレンジャーのスケジュールをパース。"""
    soup = BeautifulSoup(html, "html.parser")
    showings = []
    text = soup.get_text(separator="\n")
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    # 構造: 作品名 → (数行) → 開始時刻 → 終了時刻
    current_title = ""
    for i, line in enumerate(lines):
        # 時刻行: "9:50" or "18:10" (単独の時刻)
        m = re.match(r'^(\d{1,2}):(\d{2})$', line)
        if m:
            h, mi = int(m.group(1)), int(m.group(2))
            # 次の行が終了時刻かチェック
            if (i + 1 < len(lines)
                    and re.match(r'^\d{1,2}:\d{2}$', lines[i + 1])):
                start = f"{h:02d}:{mi:02d}"
                if current_title:
                    # 既存の作品に追加 or 新規
                    found = False
                    for s in showings:
                        if s["title"] == current_title:
                            if start not in s["times"]:
                                s["times"].append(start)
                            found = True
                            break
                    if not found:
                        showings.append({
                            "title": current_title,
                            "times": [start],
                            "description": "",
                        })
        elif (len(line) > 2
              and not re.match(r'^\d', line)
              and "もっとみる" not in line
              and "座席表" not in line
              and "スクリーン" not in line
              and "購入" not in line
              and "字幕" not in line
              and "2D" not in line
              and "ニュース" not in line
              and "上映スケジュール" not in line
              and "上映中" not in line
              and "上映履歴" not in line
              and "劇場案内" not in line
              and "アクセス" not in line
              and "ONLINE" not in line
              and "CAFE" not in line
              and "Stranger" not in line):
            candidate = re.sub(r'\s*(PG12|G|R15|R18)\s*$', '', line)
            if len(candidate.strip()) > 1:
                current_title = candidate.strip()
    return showings


def parse_musashino(html):
    """武蔵野館（cineticket.jp）のスケジュールをパース。"""
    soup = BeautifulSoup(html, "html.parser")
    showings = []
    text = soup.get_text(separator="\n")
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    current_title = ""
    # 当日の日付文字列を生成
    today_str = f"{TODAY.month:02d}月{TODAY.day:02d}日"
    in_today_block = False
    passed_header = False
    for i, line in enumerate(lines):
        # 日付タブ行をスキップ（ヘッダー部分）
        if re.match(r'^\d{2}月\d{2}日', line) and not passed_header:
            if "劇場" in lines[i + 1] if i + 1 < len(lines) else False:
                passed_header = True
                if today_str in line:
                    in_today_block = True
                continue
            continue
        # 本文中の日付ヘッダー
        if re.match(r'^\d{2}月\d{2}日', line) and passed_header:
            if today_str in line:
                in_today_block = True
            else:
                if in_today_block:
                    break  # 当日ブロック終了
            continue
        if not in_today_block:
            # "スケジュールをまとめて閉じる" の後が本文開始
            if "スケジュールをまとめて閉じる" in line:
                in_today_block = True
            continue
        # 時刻行: "12:25" (半角)
        m = re.match(r'^(\d{1,2}):(\d{2})$', line)
        if m:
            h, mi = int(m.group(1)), int(m.group(2))
            # 次の行が "～" で、その次が終了時刻か
            if (i + 2 < len(lines)
                    and lines[i + 1] == "～"
                    and re.match(r'^\d{1,2}:\d{2}$', lines[i + 2])):
                start = f"{h:02d}:{mi:02d}"
                if current_title:
                    found = False
                    for s in showings:
                        if s["title"] == current_title:
                            if start not in s["times"]:
                                s["times"].append(start)
                            found = True
                            break
                    if not found:
                        showings.append({
                            "title": current_title,
                            "times": [start],
                            "description": "",
                        })
        elif (len(line) > 2
              and not re.match(r'^[\d◎※￥¥]', line)
              and "ｽｸﾘｰﾝ" not in line
              and "詳細情報" not in line
              and "上映時間" not in line
              and "終了" not in line[:3]
              and "劇場" != line
              and "～" != line
              and "スケジュール" not in line):
            candidate = re.sub(r'\s+[GR]\s*$', '', line).strip()
            candidate = re.sub(r'\s*(PG12|R15\+?)\s*$', '', candidate)
            if len(candidate) > 2:
                current_title = candidate
    return showings


def parse_eigacom(html, target_date=None):
    """映画.comの劇場ページをDOMベースでパース。section単位で正確に抽出。"""
    soup = BeautifulSoup(html, "html.parser")
    showings = []
    d = target_date or TODAY
    today_label = f"{d.month}/{d.day}"
    # 作品ブロック: <section id="mXXXXX">
    sections = soup.find_all("section", id=re.compile(r"^m\d+"))
    for sec in sections:
        h2 = sec.find("h2", class_="title-xlarge")
        if not h2:
            continue
        title = h2.get_text(strip=True)
        # 作品URL
        movie_link = sec.find("a", href=re.compile(r"/movie/\d+"))
        movie_url = ""
        if movie_link:
            href = movie_link["href"]
            if "/movie-theater/" not in href:
                movie_url = "https://eiga.com" + href
        # 当日の時刻を<td>から抽出
        start_times = []
        for td in sec.find_all("td"):
            txt = td.get_text(strip=True)
            if not txt.startswith(today_label):
                continue
            # "3/31（火）10:4516:4021:00～22:44"
            # ～以降を除去して開始時刻のみ取得
            parts = txt.split("～")
            before_end = parts[0] if parts else txt
            times = re.findall(r"(\d{1,2}:\d{2})", before_end)
            for t in times:
                h, m = t.split(":")
                fmt = f"{int(h):02d}:{m}"
                if fmt not in start_times:
                    start_times.append(fmt)
        if start_times:
            showings.append({
                "title": title,
                "times": start_times,
                "description": "",
                "movie_url": movie_url,
            })
    return showings


def fetch_movie_description(movie_url):
    """eiga.comの作品ページから解説・あらすじを取得。"""
    try:
        resp = requests.get(movie_url, headers=HEADERS, timeout=10)
        resp.encoding = "utf-8"
        if resp.status_code != 200:
            return ""
        soup = BeautifulSoup(resp.text, "html.parser")
        # metaタグのdescriptionから取得
        meta = soup.find("meta", attrs={"name": "description"})
        if not meta or not meta.get("content"):
            return ""
        raw = meta["content"].strip()
        # 定型文「〜の作品情報。上映スケジュール、...予告動画。」を除去
        cut = re.split(
            r'の作品情報。上映スケジュール[^。]*。', raw, maxsplit=1
        )
        desc = cut[-1].strip() if len(cut) > 1 else raw
        # 末尾の製作情報を除去（「XXXX年製作／」以降）
        desc = re.split(r'\d{4}年製作', desc)[0].strip()
        return desc if desc else ""
    except Exception:
        return ""


PARSER_MAP = {
    "shimotakaido": parse_shimotakaido,
    "waseda": parse_waseda,
    "jackandbetty": parse_jackandbetty,
    "stranger": parse_stranger,
    "musashino": parse_musashino,
    "eigacom": parse_eigacom,
    "imageforum": parse_generic,
    "meguro": parse_generic,
    "kineca": parse_generic,
    "eurospace": parse_generic,
    "kscinema": parse_generic,
}


# --- メイン処理 ---

def collect_all_schedules():
    """全映画館のスケジュールをeiga.comから収集（複数日対応）。"""
    all_data = {}  # {date: [items]}
    for d in TARGET_DATES:
        all_data[d] = []

    print(f"\n[{TODAY.strftime('%Y-%m-%d')}] schedule shutoku kaishi")
    print(f"  target dates: {[d.strftime('%m/%d') for d in TARGET_DATES]}\n")

    # 作品概要キャッシュ
    desc_cache = {}

    for theater_name, info in THEATERS.items():
        print(f"  [{theater_name}] eiga.com kara shutoku...")
        url = info["eigacom_url"]
        html = fetch_html_js(url, wait_ms=5000)
        if not html:
            for d in TARGET_DATES:
                all_data[d].append({
                    "theater": theater_name,
                    "station": info["station"],
                    "address": info["address"],
                    "lat": info.get("lat"), "lng": info.get("lng"),
                    "url": info["url_top"],
                    "eigacom_url": url,
                    "title": "(shutoku shippai)",
                    "time": "", "description": "", "movie_url": "",
                })
            continue

        for d in TARGET_DATES:
            is_wd = d.weekday() < 5 and d not in HOLIDAYS_2026
            showings = parse_eigacom(html, target_date=d)
            if is_wd:
                showings = filter_by_time(showings)

            if not showings:
                all_data[d].append({
                    "theater": theater_name,
                    "station": info["station"],
                    "address": info["address"],
                    "lat": info.get("lat"), "lng": info.get("lng"),
                    "url": info["url_top"],
                    "eigacom_url": url,
                    "title": "(gaitou jikantai no jouei nashi)" if is_wd else "(jouei nashi)",
                    "time": "", "description": "", "movie_url": "",
                })
                continue

            for s in showings:
                time_str = ", ".join(s.get("times", []))
                desc = ""
                movie_url = s.get("movie_url", "")
                if movie_url:
                    if movie_url in desc_cache:
                        desc = desc_cache[movie_url]
                    else:
                        desc = fetch_movie_description(movie_url)
                        desc_cache[movie_url] = desc
                all_data[d].append({
                    "theater": theater_name,
                    "station": info["station"],
                    "address": info["address"],
                    "lat": info.get("lat"), "lng": info.get("lng"),
                    "url": info["url_top"],
                    "eigacom_url": url,
                    "title": s["title"],
                    "time": time_str,
                    "description": desc,
                    "movie_url": movie_url,
                })
        count = sum(1 for d in TARGET_DATES
                    for item in all_data[d]
                    if item["theater"] == theater_name
                    and "shippai" not in item["title"]
                    and "jouei" not in item["title"]
                    and "nashi" not in item["title"])
        if count > 0:
            print(f"    -> {count} ken shutoku")

    return all_data


def save_to_excel(data):
    """結果をExcelファイルに保存。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("[WARN] openpyxlが必要です: pip install openpyxl")
        save_to_csv(data)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    day_type = "平日" if IS_WEEKDAY else "休日"
    ws.title = f"上映スケジュール_{TODAY.strftime('%m%d')}"

    # ヘッダー行
    headers = ["映画館", "最寄り駅", "住所", "公式サイト", "上映開始時刻", "作品名", "作品概要"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # データ行
    # 時刻でソート
    data_sorted = sorted(data, key=lambda x: (x["theater"], x["time"]))
    for row_idx, item in enumerate(data_sorted, 2):
        values = [
            item["theater"],
            item["station"],
            item["address"],
            item["url"],
            item["time"],
            item["title"],
            item["description"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 列幅調整
    col_widths = [22, 30, 35, 35, 18, 30, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # タイトル行を固定
    ws.freeze_panes = "A2"

    # 保存
    filename = f"ミニシアター上映スケジュール_{TODAY.strftime('%Y%m%d')}_{day_type}.xlsx"
    filepath = os.path.join(os.path.dirname(__file__), filename)
    wb.save(filepath)
    print(f"\n[OK] 保存完了: {filepath}")
    return filepath


def save_to_csv(data):
    """openpyxlがない場合のフォールバック。"""
    import csv
    day_type = "平日" if IS_WEEKDAY else "休日"
    filename = f"ミニシアター上映スケジュール_{TODAY.strftime('%Y%m%d')}_{day_type}.csv"
    filepath = os.path.join(os.path.dirname(__file__), filename)
    with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "theater", "station", "address", "url", "time", "title", "description"
        ])
        writer.writeheader()
        writer.writerows(data)
    print(f"\n[OK] CSV保存完了: {filepath}")
    return filepath



def save_to_html(data):
    """複数日タブ対応のレスポンシブHTML生成。"""
    weekday_jp = "月火水木金土日"
    date_str = TODAY.strftime('%Y年%m月%d日')

    area_chips_html = build_area_chips(data.values())

    # 1日分のみの場合はタブ不要（ただしJS処理を揃えるため.tab-panelで統一）
    if len(TARGET_DATES) == 1:
        d = TARGET_DATES[0]
        is_wd = d.weekday() < 5 and d not in HOLIDAYS_2026
        d_type = "平日" if is_wd else "休日"
        time_filter = "18-22時" if is_wd else "終日"
        items = data[d]
        panel_html = build_day_panel(items, d_type, time_filter)
        d_label = f"{d.month}/{d.day}({weekday_jp[d.weekday()]})"
        tab_html = (
            f'<div class="tab-panel" id="p0" data-date="{d.isoformat()}"'
            f' data-date-label="{html.escape(d_label)}" style="display:block">'
            f'{panel_html}</div>\n'
        )
        html_content = build_full_html(
            date_str, area_chips_html, "", tab_html, ""
        )
    else:
        # 複数日: radioタブで切り替え
        # ラベルを先に全て並べ、パネルを後に並べる
        tab_inputs_labels = ""
        tab_panels = ""
        for idx, d in enumerate(TARGET_DATES):
            d_label = f"{d.month}/{d.day}({weekday_jp[d.weekday()]})"
            d_id = f"tab{idx}"
            checked = " checked" if idx == 0 else ""
            is_wd = d.weekday() < 5 and d not in HOLIDAYS_2026
            d_type = "平日" if is_wd else "休日"
            time_filter = "18-22時" if is_wd else "終日"
            items = data[d]
            id_prefix = f"{d_id}-"
            panel_html = build_day_panel(items, d_type, time_filter, id_prefix)
            tab_inputs_labels += (
                f'<input type="radio" name="tabs" id="{d_id}"'
                f'{checked} class="tab-input">\n'
                f'<label for="{d_id}" class="tab-label">'
                f'{d_label}</label>\n'
            )
            tab_panels += (
                f'<div class="tab-panel" id="p{idx}" data-date="{d.isoformat()}"'
                f' data-date-label="{html.escape(d_label)}">'
                f'{panel_html}</div>\n'
            )
        tab_html = tab_inputs_labels + tab_panels
        html_content = build_full_html(
            date_str, area_chips_html, "", tab_html, ""
        )

    filename = "index.html"
    box_dir = r"C:\Users\gokawa\Box\Go-CINEEMA-app"
    local_dir = os.path.dirname(__file__)
    local_path = os.path.join(local_dir, filename)
    with open(local_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"[OK] HTML: {local_path}")
    if os.path.isdir(box_dir):
        import shutil
        box_path = os.path.join(box_dir, filename)
        shutil.copy2(local_path, box_path)
        print(f"[OK] BOX: {box_path}")
    return local_path


# --- HTMLヘルパー関数 ---

AREA_RULES = [
    ("新宿", ["新宿駅", "新宿三丁目"]),
    ("有楽町・日比谷・銀座", ["有楽町", "日比谷", "銀座", "京橋"]),
    ("渋谷・道玄坂", ["渋谷駅", "道玄坂", "ハチ公"]),
    ("目黒・恵比寿", ["目黒駅", "恵比寿"]),
    ("池袋・大塚", ["池袋", "大塚"]),
    ("神保町・御茶ノ水", ["神保町", "御茶ノ水"]),
    ("下北沢", ["下北沢"]),
    ("高田馬場", ["高田馬場"]),
    ("飯田橋", ["飯田橋"]),
    ("下高井戸", ["下高井戸"]),
    ("東中野", ["東中野"]),
    ("大森", ["大森"]),
    ("菊川", ["菊川"]),
    ("田端", ["田端"]),
    ("北千住", ["北千住"]),
    ("品川・大井町", ["品川", "大井町"]),
    ("横浜", ["黄金町"]),
]
AREA_ORDER = [a[0] for a in AREA_RULES] + ["その他"]


def get_area(station):
    """最寄り駅からエリアを判定。"""
    for area_name, keywords in AREA_RULES:
        for kw in keywords:
            if kw in station:
                return area_name
    return "その他"


def movie_id(m):
    """作品を一意に識別するキー（お気に入り等に使用）。"""
    movie_url = m.get("movie_url", "")
    mo = re.search(r"/movie/(\d+)", movie_url)
    if mo:
        return f"mv{mo.group(1)}"
    return "t-" + re.sub(r"[^0-9A-Za-z぀-ヿ一-鿿]+", "-", m["title"])


def build_movie_html(m, theater_name=""):
    """1作品分のHTMLを生成。"""
    title = m["title"]
    time_str = m["time"] if m["time"] else ""
    desc = m.get("description") or ""
    movie_url = m.get("movie_url", "")
    is_fb = any(k in title for k in [
        "shutoku", "shippai", "gaitou", "jouei", "nashi", "fuka"
    ])
    if is_fb:
        return ""
    title_esc = html.escape(title)
    title_link = (
        f'<a href="{html.escape(movie_url)}" target="_blank"'
        f' class="mv-title-link">{title_esc}</a>'
        if movie_url else title_esc
    )
    desc_html = ""
    if desc:
        short = desc[:50] + "..." if len(desc) > 50 else desc
        if len(desc) > 50:
            desc_html = (
                f'<details class="mv-details">'
                f'<summary class="mv-desc-short">{html.escape(short)}</summary>'
                f'<p class="mv-desc-full">{html.escape(desc)}</p></details>'
            )
        else:
            desc_html = f'<p class="mv-desc-short">{html.escape(desc)}</p>'
    times = [t for t in time_str.split(", ") if t]
    badges = "".join(f'<span class="badge">{t}</span>' for t in times)
    mid = movie_id(m)
    times_attr = html.escape(" ".join(times))
    return f"""
      <div class="mv" data-movie-id="{mid}" data-title="{title_esc.lower()}" data-times="{times_attr}" data-theater="{html.escape(theater_name)}">
        <div class="mv-head">
          <div class="mv-times">{badges}</div>
          <div class="mv-title-row">
            <h3 class="mv-title">{title_link}</h3>
            <button class="fav-btn" data-fav-id="{mid}" data-fav-title="{title_esc}" aria-label="この作品をお気に入りに登録">&#9734;</button>
          </div>
        </div>
        {desc_html}
      </div>"""


def build_day_panel(items, day_type, time_filter, id_prefix=""):
    """1日分のエリア別映画館カードHTMLを生成。"""
    grouped = {}
    for item in items:
        name = item["theater"]
        if name not in grouped:
            grouped[name] = {
                "station": item["station"],
                "address": item["address"],
                "lat": item.get("lat"),
                "lng": item.get("lng"),
                "url": item["url"],
                "eigacom_url": item.get("eigacom_url", ""),
                "movies": [],
            }
        grouped[name]["movies"].append(item)

    area_theaters = {}
    for th_name, info in grouped.items():
        area = get_area(info["station"])
        if area not in area_theaters:
            area_theaters[area] = []
        area_theaters[area].append((th_name, info))

    out = f'<div class="day-info">{day_type} / {time_filter}</div>'
    for area_name in AREA_ORDER:
        if area_name not in area_theaters:
            continue
        out += f'<div class="area-header" id="area-{id_prefix}{area_name}" data-area="{html.escape(area_name)}">{area_name}</div>'
        for th_name, info in area_theaters[area_name]:
            movies = "".join(
                build_movie_html(m, th_name)
                for m in sorted(info["movies"], key=lambda x: x["time"])
            )
            if not movies.strip():
                continue
            url = info["url"]
            eigacom = info.get("eigacom_url", "")
            links = f'<a href="{url}" target="_blank">公式</a>'
            if eigacom:
                links += f' <a href="{eigacom}" target="_blank">eiga.com</a>'
            lat = info.get("lat")
            lng = info.get("lng")
            geo_attr = (
                f' data-lat="{lat}" data-lng="{lng}"' if lat and lng else ""
            )
            out += f"""
    <section class="theater" data-area="{html.escape(area_name)}"{geo_attr}>
      <div class="th-head">
        <h2>{html.escape(th_name)}</h2>
        <div class="th-meta">
          <span>{ICON_TRAIN} {html.escape(info["station"])}</span>
          <span>{ICON_PIN} {html.escape(info["address"])}</span>
          <span class="th-dist" hidden></span>
        </div>
        <div class="th-links">{links}</div>
      </div>
      <div class="mv-list">{movies}</div>
    </section>"""
    return out




def build_area_chips(all_items):
    """エリア複数選択フィルター用のチップHTMLを生成（全日程分のエリア集合）。"""
    areas_present = set()
    for items in all_items:
        for item in items:
            areas_present.add(get_area(item["station"]))
    return "".join(
        f'<button class="chip" data-area="{html.escape(a)}">{a}</button>'
        for a in AREA_ORDER if a in areas_present
    )


ICON_PIN = (
    '<svg viewBox="0 0 24 24" width="12" height="12" fill="none" '
    'stroke="currentColor" stroke-width="2" stroke-linecap="round" '
    'stroke-linejoin="round" style="vertical-align:-1px"><path '
    'd="M20 10c0 6-8 12-8 12s-8-6-8-12a8 8 0 1 1 16 0Z"/>'
    '<circle cx="12" cy="10" r="2.5"/></svg>'
)
ICON_TRAIN = (
    '<svg viewBox="0 0 24 24" width="12" height="12" fill="none" '
    'stroke="currentColor" stroke-width="2" stroke-linecap="round" '
    'stroke-linejoin="round" style="vertical-align:-1px"><rect x="4" y="3" '
    'width="16" height="14" rx="3"/><path d="M4 11h16M9 17l-2 3M15 17l2 3"/>'
    '<circle cx="8.5" cy="7" r="0.5" fill="currentColor"/>'
    '<circle cx="15.5" cy="7" r="0.5" fill="currentColor"/></svg>'
)
ICON_LOCATE = (
    '<svg viewBox="0 0 24 24" width="14" height="14" fill="none" '
    'stroke="currentColor" stroke-width="2" stroke-linecap="round" '
    'stroke-linejoin="round"><circle cx="12" cy="12" r="3"/>'
    '<path d="M12 2v3M12 19v3M2 12h3M19 12h3"/></svg>'
)
ICON_BOOKMARK = (
    '<svg viewBox="0 0 24 24" width="14" height="14" fill="none" '
    'stroke="currentColor" stroke-width="2" stroke-linecap="round" '
    'stroke-linejoin="round"><path d="M6 3h12v18l-6-4-6 4V3Z"/></svg>'
)

MAIN_JS = r"""
const FAV_KEY = 'mts_favorites_v1';

function loadFavs() {
  try { return JSON.parse(localStorage.getItem(FAV_KEY) || '{}'); }
  catch (e) { return {}; }
}
function saveFavs(favs) { localStorage.setItem(FAV_KEY, JSON.stringify(favs)); }

function refreshFavButtons() {
  const favs = loadFavs();
  document.querySelectorAll('.fav-btn').forEach(btn => {
    const active = !!favs[btn.dataset.favId];
    btn.classList.toggle('active', active);
    btn.textContent = active ? '★' : '☆';
  });
}

function escapeHtml(s) {
  const d = document.createElement('div');
  d.textContent = s;
  return d.innerHTML;
}

function getActivePanel() {
  const panels = document.querySelectorAll('.tab-panel');
  const checked = document.querySelector('.tab-input:checked');
  if (!checked) return panels[0];
  const inputs = Array.from(document.querySelectorAll('.tab-input'));
  return panels[inputs.indexOf(checked)] || panels[0];
}

function applyFilters() {
  const searchVal = document.getElementById('searchBox').value.trim().toLowerCase();
  const activeAreas = Array.from(document.querySelectorAll('.chip[data-area].active')).map(c => c.dataset.area);

  document.querySelectorAll('.tab-panel').forEach(panel => {
    let currentHeader = null;
    let currentAreaVisible = false;
    Array.from(panel.children).forEach(el => {
      if (el.classList.contains('area-header')) {
        if (currentHeader) currentHeader.style.display = currentAreaVisible ? '' : 'none';
        currentHeader = el;
        currentAreaVisible = false;
        return;
      }
      if (!el.classList.contains('theater')) return;
      const areaMatch = activeAreas.length === 0 || activeAreas.includes(el.dataset.area);
      let theaterVisible = false;
      el.querySelectorAll('.mv').forEach(mv => {
        const title = mv.dataset.title || '';
        const searchMatch = !searchVal || title.includes(searchVal);
        const visible = searchMatch && areaMatch;
        mv.classList.toggle('mv-hidden', !visible);
        if (visible) theaterVisible = true;
      });
      el.style.display = theaterVisible ? '' : 'none';
      if (theaterVisible) currentAreaVisible = true;
    });
    if (currentHeader) currentHeader.style.display = currentAreaVisible ? '' : 'none';
  });

  const anyFilter = !!(searchVal || activeAreas.length);
  const statusEl = document.getElementById('filterStatus');
  statusEl.hidden = !anyFilter;
  if (anyFilter) {
    const activePanel = getActivePanel();
    const visible = activePanel.querySelectorAll('.mv:not(.mv-hidden)').length;
    const all = activePanel.querySelectorAll('.mv').length;
    document.getElementById('filterCount').textContent = visible + '件 / 全' + all + '件';
  }
}

function renderMylist() {
  const favs = loadFavs();
  const body = document.getElementById('mylistBody');
  const ids = Object.keys(favs);
  if (ids.length === 0) {
    body.innerHTML = '<p class="mylist-empty">★パタンで気になる作品を登録すると、上映館と時間がここに表示されます。</p>';
    return;
  }
  const showingsByMovie = {};
  document.querySelectorAll('.tab-panel').forEach(panel => {
    const dateLabel = panel.dataset.dateLabel || '';
    panel.querySelectorAll('.mv[data-movie-id]').forEach(mv => {
      const id = mv.dataset.movieId;
      if (!favs[id]) return;
      const theater = mv.dataset.theater || '';
      const times = (mv.dataset.times || '').split(' ').filter(Boolean);
      if (!showingsByMovie[id]) showingsByMovie[id] = [];
      times.forEach(t => showingsByMovie[id].push({ dateLabel: dateLabel, theater: theater, time: t }));
    });
  });
  let out = '';
  ids.forEach(id => {
    const title = (favs[id] && favs[id].title) || '(不明な作品)';
    out += '<div class="mylist-item"><h3>' + escapeHtml(title) + '</h3>';
    const showings = showingsByMovie[id];
    if (showings && showings.length) {
      showings.forEach(s => {
        out += '<div class="mylist-showing"><span class="mylist-date">' + escapeHtml(s.dateLabel) +
               '</span> <span class="mylist-theater">' + escapeHtml(s.theater) + '</span> ' + escapeHtml(s.time) + '</div>';
      });
    } else {
      out += '<div class="mylist-none">現在の上映スケジュールには見つかりませんでした</div>';
    }
    out += '</div>';
  });
  body.innerHTML = out;
}

function haversineKm(lat1, lon1, lat2, lon2) {
  const R = 6371;
  const dLat = (lat2 - lat1) * Math.PI / 180;
  const dLon = (lon2 - lon1) * Math.PI / 180;
  const a = Math.sin(dLat / 2) ** 2 +
            Math.cos(lat1 * Math.PI / 180) * Math.cos(lat2 * Math.PI / 180) * Math.sin(dLon / 2) ** 2;
  return R * 2 * Math.atan2(Math.sqrt(a), Math.sqrt(1 - a));
}

let panelOriginalOrder = null;
function captureOriginalOrder() {
  panelOriginalOrder = new Map();
  document.querySelectorAll('.tab-panel').forEach(panel => {
    panelOriginalOrder.set(panel, Array.from(panel.children));
  });
}

function sortByDistance(lat, lon) {
  document.querySelectorAll('.tab-panel').forEach(panel => {
    panel.querySelectorAll('.area-header').forEach(h => { h.style.display = 'none'; });
    const theaters = Array.from(panel.querySelectorAll('.theater[data-lat][data-lng]'));
    theaters.forEach(t => {
      const d = haversineKm(lat, lon, parseFloat(t.dataset.lat), parseFloat(t.dataset.lng));
      t.dataset.distance = d.toFixed(1);
      const distEl = t.querySelector('.th-dist');
      if (distEl) {
        distEl.innerHTML = DIST_ICON_HTML + ' 現在地から ' + d.toFixed(1) + 'km';
        distEl.hidden = false;
      }
    });
    theaters.sort((a, b) => parseFloat(a.dataset.distance) - parseFloat(b.dataset.distance));
    theaters.forEach(t => panel.appendChild(t));
  });
}

function restoreOriginalOrder() {
  if (!panelOriginalOrder) return;
  panelOriginalOrder.forEach((children, panel) => {
    children.forEach(c => panel.appendChild(c));
  });
  document.querySelectorAll('.th-dist').forEach(d => { d.hidden = true; d.textContent = ''; });
  applyFilters();
}

document.addEventListener('DOMContentLoaded', () => {
  refreshFavButtons();
  captureOriginalOrder();
  applyFilters();

  document.getElementById('searchBox').addEventListener('input', applyFilters);
  document.querySelectorAll('.tab-input').forEach(inp => inp.addEventListener('change', applyFilters));

  document.querySelectorAll('.chip[data-area]').forEach(chip => {
    chip.addEventListener('click', () => {
      const wasActive = chip.classList.contains('active');
      document.querySelectorAll('.chip[data-area].active').forEach(c => c.classList.remove('active'));
      if (!wasActive) {
        chip.classList.add('active');
        applyFilters();
        const panel = getActivePanel();
        const header = panel.querySelector('.area-header[data-area="' + chip.dataset.area + '"]');
        if (header) header.scrollIntoView({ behavior: 'smooth', block: 'start' });
      } else {
        applyFilters();
      }
    });
  });

  document.getElementById('clearFiltersBtn').addEventListener('click', () => {
    document.getElementById('searchBox').value = '';
    document.querySelectorAll('.chip.active').forEach(c => c.classList.remove('active'));
    applyFilters();
  });

  document.addEventListener('click', e => {
    const btn = e.target.closest('.fav-btn');
    if (!btn) return;
    const id = btn.dataset.favId;
    const favs = loadFavs();
    if (favs[id]) {
      delete favs[id];
    } else {
      favs[id] = { title: btn.dataset.favTitle, savedAt: new Date().toISOString() };
    }
    saveFavs(favs);
    refreshFavButtons();
    if (document.getElementById('mylistOverlay').classList.contains('open')) renderMylist();
  });

  document.getElementById('mylistBtn').addEventListener('click', () => {
    renderMylist();
    document.getElementById('mylistOverlay').classList.add('open');
  });
  document.getElementById('mylistClose').addEventListener('click', () => {
    document.getElementById('mylistOverlay').classList.remove('open');
  });
  document.getElementById('mylistOverlay').addEventListener('click', e => {
    if (e.target.id === 'mylistOverlay') e.target.classList.remove('open');
  });

  document.getElementById('nearbyBtn').addEventListener('click', () => {
    const btn = document.getElementById('nearbyBtn');
    if (btn.classList.contains('active')) {
      btn.classList.remove('active');
      restoreOriginalOrder();
      return;
    }
    if (!navigator.geolocation) {
      alert('この端末では位置情報を利用できません');
      return;
    }
    navigator.geolocation.getCurrentPosition(
      pos => {
        sortByDistance(pos.coords.latitude, pos.coords.longitude);
        btn.classList.add('active');
      },
      () => {
        alert('位置情報を取得できませんでした。ブラウザの位置情報設定をご確認ください。');
      },
      { timeout: 8000 }
    );
  });
});
""".strip().replace("DIST_ICON_HTML", json.dumps(ICON_LOCATE))


def build_full_html(date_str, area_chips_html, _unused2,
                    tab_panels, nav_links):
    """完全なHTMLドキュメントを生成。"""
    return f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Mini Theater Schedule - {date_str}</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Zen+Kaku+Gothic+New:wght@400;500;700;900&family=Zen+Old+Mincho:wght@600&display=swap" rel="stylesheet">
<style>
:root {{
  --bg: #0a0a0c; --bg-soft: #0f0f12;
  --card: #141417; --card-head: #191a1e;
  --border: #232327; --border-soft: #1b1b1f;
  --text: #ececec; --sub: #85818a; --faint: #52505a;
  --gold: #d4a24c; --gold-deep: #8f6c2e; --gold-glow: rgba(212,162,76,.14);
  --link: #9db4d6;
  --serif: 'Zen Old Mincho', serif;
  --sans: 'Zen Kaku Gothic New', -apple-system, BlinkMacSystemFont,
          'Hiragino Sans', 'Noto Sans JP', 'Segoe UI', sans-serif;
}}
* {{ margin:0; padding:0; box-sizing:border-box; }}
html {{ background: var(--bg); }}
body {{
  font-family: var(--sans);
  background: var(--bg); color: var(--text);
  max-width: 720px; margin: 0 auto; padding: 0 0 40px;
  -webkit-font-smoothing: antialiased;
  letter-spacing: .01em;
}}
a {{ -webkit-tap-highlight-color: transparent; }}
/* --- Header --- */
.hdr {{
  position: sticky; top: 0; z-index: 20;
  background: linear-gradient(180deg, var(--bg) 82%, transparent);
  backdrop-filter: blur(6px);
  padding: 22px 16px 12px;
}}
.hdr-top {{
  display: flex; align-items: baseline; justify-content: space-between;
  padding-bottom: 14px; border-bottom: 1px solid var(--border-soft);
}}
.hdr-brand {{ display: flex; align-items: baseline; gap: 10px; }}
.hdr h1 {{
  font-family: var(--serif); font-size: 19px; font-weight: 600;
  letter-spacing: .06em; color: var(--text);
}}
.hdr .kicker {{
  font-size: 9px; font-weight: 700; letter-spacing: .22em;
  color: var(--gold); text-transform: uppercase;
}}
.hdr .date {{
  font-size: 11px; color: var(--sub); letter-spacing: .05em;
  font-variant-numeric: tabular-nums;
}}
/* --- Toolbar --- */
.toolbar {{ padding-top: 14px; }}
.toolbar-row {{ display: flex; gap: 8px; }}
.search-box {{
  flex: 1; min-width: 0; background: var(--card); color: var(--text);
  border: 1px solid var(--border); border-radius: 12px;
  padding: 10px 14px; font-size: 13px; font-family: var(--sans);
  transition: border-color .15s;
}}
.search-box:focus {{ outline: none; border-color: var(--gold-deep); }}
.search-box::placeholder {{ color: var(--faint); }}
.tool-btn {{
  font-size: 11px; font-weight: 700; color: var(--sub);
  background: var(--card); border: 1px solid var(--border);
  border-radius: 12px; padding: 10px 13px; cursor: pointer;
  white-space: nowrap; letter-spacing: .03em;
  transition: color .15s, border-color .15s, background .15s;
}}
.tool-btn:hover {{ color: var(--text); border-color: var(--faint); }}
.tool-btn.active {{ color: var(--gold); border-color: var(--gold-deep); background: var(--gold-glow); }}
.chip-row {{
  display: flex; gap: 7px; margin-top: 10px;
  overflow-x: auto; scrollbar-width: none;
  padding-bottom: 2px;
  mask-image: linear-gradient(90deg, #000 92%, transparent);
}}
.chip-row::-webkit-scrollbar {{ display: none; }}
.chip {{
  flex: 0 0 auto;
  font-size: 11px; font-weight: 600; color: var(--sub);
  background: transparent; border: 1px solid var(--border);
  border-radius: 20px; padding: 5px 13px; cursor: pointer;
  white-space: nowrap; transition: color .15s, border-color .15s, background .15s;
}}
.chip:hover {{ border-color: var(--faint); color: var(--text); }}
.chip.active {{ color: var(--gold); border-color: var(--gold-deep); background: var(--gold-glow); }}
.filter-status[hidden] {{ display: none; }}
.filter-status {{
  display: flex; align-items: center; justify-content: center;
  gap: 12px; font-size: 11px; color: var(--sub); margin-top: 10px;
  letter-spacing: .03em;
}}
.filter-status button {{
  font-size: 11px; color: var(--gold); background: none; border: none;
  cursor: pointer; font-weight: 600;
}}
/* --- Tabs --- */
.tab-bar {{ display: flex; gap: 8px; padding: 4px 16px 0; flex-wrap: wrap; }}
.tab-input {{ display: none; }}
.tab-label {{
  font-size: 11px; font-weight: 700; padding: 7px 16px;
  border: 1px solid var(--border); border-radius: 20px;
  color: var(--sub); cursor: pointer; letter-spacing: .04em;
  transition: color .15s, border-color .15s, background .15s;
}}
.tab-input:checked + .tab-label {{
  color: #0a0a0c; border-color: var(--gold);
  background: var(--gold);
}}
.tab-panel {{ display: none; }}
.tab-input:nth-of-type(1):checked ~ #p0 {{ display: block; }}
.tab-input:nth-of-type(2):checked ~ #p1 {{ display: block; }}
.tab-input:nth-of-type(3):checked ~ #p2 {{ display: block; }}
.tab-input:nth-of-type(4):checked ~ #p3 {{ display: block; }}
.day-info {{
  text-align: center; font-size: 10px; color: var(--faint);
  letter-spacing: .15em; text-transform: uppercase;
  padding: 16px 0 4px;
}}
.area-header {{
  display: flex; align-items: baseline; gap: 10px;
  font-family: var(--serif); font-size: 13px; font-weight: 600; color: var(--gold);
  padding: 0 16px; margin-top: 32px; margin-bottom: 14px;
  letter-spacing: .1em;
}}
.area-header::after {{
  content: ""; flex: 1; height: 1px;
  background: linear-gradient(90deg, var(--border), transparent);
}}
.theater {{
  background: var(--card);
  margin: 0 16px 14px; overflow: hidden;
  border: 1px solid var(--border);
  border-radius: 2px;
  transition: border-color .2s;
}}
.th-head {{
  padding: 16px 18px 12px; background: var(--card-head);
  border-bottom: 1px solid var(--border-soft);
  position: relative;
}}
.th-head::before {{
  content: ""; position: absolute; left: 0; top: 0; bottom: 0; width: 2px;
  background: var(--gold-deep);
}}
.th-head h2 {{ font-size: 14px; font-weight: 700; letter-spacing: .02em; }}
.th-meta {{
  display: flex; flex-direction: column; gap: 3px;
  font-size: 10.5px; color: var(--sub); margin-top: 7px;
  letter-spacing: .01em;
}}
.th-dist {{ color: var(--gold); font-weight: 700; letter-spacing: .04em; }}
.th-links {{ margin-top: 8px; display: flex; gap: 16px; }}
.th-links a {{
  font-size: 10.5px; color: var(--link);
  text-decoration: none; font-weight: 500; letter-spacing: .02em;
}}
.th-links a:hover {{ color: var(--gold); }}
.mv-list {{ padding: 4px 18px 6px; }}
.mv {{
  padding: 14px 0; border-bottom: 1px solid var(--border-soft);
}}
.mv.mv-hidden {{ display: none; }}
.mv:last-child {{ border-bottom: none; }}
.mv-head {{ display: flex; flex-direction: column; gap: 6px; }}
.mv-times {{ display: flex; flex-wrap: wrap; gap: 6px; }}
.badge {{
  display: inline-block; font-size: 12px; font-weight: 700;
  color: var(--gold); background: transparent;
  border: 1px solid var(--gold-deep);
  padding: 2px 9px; border-radius: 3px;
  letter-spacing: .02em; font-variant-numeric: tabular-nums;
}}
.mv-title-row {{ display: flex; align-items: flex-start; gap: 6px; }}
.mv-title {{ flex: 1; font-size: 14px; font-weight: 600; line-height: 1.5; }}
.mv-title-link {{ color: var(--text); text-decoration: none; }}
.mv-title-link:hover {{ color: var(--gold); }}
.fav-btn {{
  flex: 0 0 auto; background: none; border: none;
  font-size: 18px; line-height: 1; color: var(--faint); cursor: pointer;
  padding: 0 2px; margin-top: 1px; transition: color .15s, transform .15s;
}}
.fav-btn:hover {{ color: var(--gold-deep); }}
.fav-btn.active {{ color: var(--gold); }}
.mv-details {{ margin-top: 7px; }}
.mv-details summary {{
  list-style: none; cursor: pointer;
  font-size: 12px; color: var(--sub); line-height: 1.7;
}}
.mv-details summary::-webkit-details-marker {{ display: none; }}
.mv-details summary::after {{ content: " 続きを読む"; font-size: 10px; color: var(--gold-deep); }}
.mv-details[open] summary {{ display: none; }}
.mv-desc-short {{
  font-size: 12px; color: var(--sub); line-height: 1.7; margin-top: 7px;
}}
.mv-desc-full {{
  font-size: 12px; color: var(--sub); line-height: 1.7; margin-top: 0;
}}
.ftr {{
  text-align: center; padding: 40px 16px 8px;
  font-size: 10px; color: var(--faint);
  letter-spacing: .04em;
}}
.ftr .ftr-rule {{
  width: 40px; height: 1px; background: var(--border);
  margin: 0 auto 16px;
}}
/* --- Mylist overlay --- */
.mylist-overlay {{
  display: none; position: fixed; inset: 0; background: rgba(6,6,8,.72);
  z-index: 100; padding: 0; backdrop-filter: blur(2px);
}}
.mylist-overlay.open {{ display: block; }}
.mylist-panel {{
  background: var(--bg); max-width: 720px; margin: 0 auto;
  height: 100%; overflow-y: auto; border-left: 1px solid var(--border);
  border-right: 1px solid var(--border);
}}
.mylist-head {{
  display: flex; align-items: center; justify-content: space-between;
  padding: 18px 20px; border-bottom: 1px solid var(--border-soft);
  position: sticky; top: 0; background: var(--bg);
}}
.mylist-head h2 {{
  display: flex; align-items: center; gap: 8px;
  font-family: var(--serif); font-size: 16px; font-weight: 600;
  letter-spacing: .06em; color: var(--gold);
}}
.mylist-close {{
  background: none; border: none; color: var(--sub); font-size: 20px;
  cursor: pointer; padding: 4px;
}}
.mylist-close:hover {{ color: var(--text); }}
.mylist-body {{ padding: 16px 20px 32px; }}
.mylist-empty {{ color: var(--sub); font-size: 13px; text-align: center; padding: 60px 20px; line-height: 1.8; }}
.mylist-item {{
  background: var(--card); border: 1px solid var(--border);
  padding: 16px 18px; margin-bottom: 12px; position: relative;
}}
.mylist-item::before {{
  content: ""; position: absolute; left: 0; top: 0; bottom: 0; width: 2px;
  background: var(--gold-deep);
}}
.mylist-item h3 {{ font-size: 14px; font-weight: 700; margin-bottom: 8px; letter-spacing: .02em; }}
.mylist-showing {{ font-size: 12px; color: var(--sub); padding: 4px 0; }}
.mylist-date {{ color: var(--gold); font-weight: 700; }}
.mylist-theater {{ color: var(--text); }}
.mylist-none {{ font-size: 12px; color: var(--faint); }}
</style>
</head>
<body>
<header class="hdr">
  <div class="hdr-top">
    <div class="hdr-brand">
      <h1>Mini Theater</h1>
      <span class="kicker">Schedule</span>
    </div>
    <div class="date">{date_str}</div>
  </div>
  <div class="toolbar">
    <div class="toolbar-row">
      <input type="search" id="searchBox" class="search-box" placeholder="作品名で検索">
      <button id="nearbyBtn" class="tool-btn" aria-label="現在地から近い順">{ICON_LOCATE}</button>
      <button id="mylistBtn" class="tool-btn" aria-label="お気に入り作品の一覧を見る">{ICON_BOOKMARK}</button>
    </div>
    <div class="chip-row" id="areaChips">{area_chips_html}</div>
    <div class="filter-status" id="filterStatus" hidden>
      <span id="filterCount"></span>
      <button id="clearFiltersBtn">クリア</button>
    </div>
  </div>
</header>
<div class="tabs">
{tab_panels}
</div>
<footer class="ftr">
  <div class="ftr-rule"></div>
  {datetime.now(JST).strftime('%H:%M')} updated &middot; source: eiga.com<br>
  schedules may change &mdash; check official sites
</footer>

<div class="mylist-overlay" id="mylistOverlay">
  <div class="mylist-panel">
    <div class="mylist-head">
      <h2>{ICON_BOOKMARK} マイリスト（お気に入り作品）</h2>
      <button id="mylistClose" class="mylist-close">&#10005;</button>
    </div>
    <div class="mylist-body" id="mylistBody"></div>
  </div>
</div>

<script>
{MAIN_JS}
</script>
</body>
</html>"""




def push_to_github():
    """GitHub APIでindex.htmlをリポジトリにpush。"""
    import base64
    token = os.environ.get("GITHUB_TOKEN", "")
    if not token:
        print("[WARN] GITHUB_TOKEN not set, skip push")
        return
    owner = "taiwan-chodofu"
    repo = "mini-theater-schedule"
    filepath = os.path.join(os.path.dirname(__file__), "index.html")
    if not os.path.exists(filepath):
        print("[WARN] index.html not found, skip push")
        return
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()
    encoded = base64.b64encode(content.encode("utf-8")).decode("ascii")
    api_url = f"https://api.github.com/repos/{owner}/{repo}/contents/index.html"
    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json",
    }
    # 既存ファイルのSHAを取得（更新時に必要）
    sha = ""
    try:
        resp = requests.get(api_url, headers=headers, timeout=10)
        if resp.status_code == 200:
            sha = resp.json().get("sha", "")
    except Exception:
        pass
    payload = {
        "message": f"update schedule {TODAY.strftime('%Y-%m-%d')}",
        "content": encoded,
    }
    if sha:
        payload["sha"] = sha
    try:
        resp = requests.put(api_url, headers=headers, json=payload, timeout=15)
        if resp.status_code in (200, 201):
            print("[OK] GitHub Pages push done")
        else:
            print(f"[WARN] GitHub push failed: {resp.status_code} {resp.text[:100]}")
    except Exception as e:
        print(f"[WARN] GitHub push error: {e}")


if __name__ == "__main__":
    data = collect_all_schedules()
    # 最初の日付分でExcel保存
    first_date = TARGET_DATES[0]
    save_to_excel(data[first_date])
    save_to_html(data)
    push_to_github()
    total = sum(len(v) for v in data.values())
    print(f"\n合計 {total} 件のスケジュールを取得しました")
    day_type = "平日" if IS_WEEKDAY else "休日"
    if IS_WEEKDAY:
        print("平日モード: 18:00-22:00開始の回のみ表示")
    else:
        print("休日モード: 終日の上映スケジュールを表示")
